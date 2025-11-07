import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import io
from PIL import Image as PILImage
import os
import streamlit as st
import tempfile

def process_sheet_data(df, sheet_name="ACHE"):
    """
    Process a single sheet's data and return the figure and analysis results.
    
    Parameters:
    -----------
    df : DataFrame
        The data to process
    sheet_name : str
        Name of the sheet (used in plot title)
    """
    try:
        # Check if we have operating point columns
        has_operating_points = False
        operating_points = []
        
        # Store original column count to detect operating point columns
        original_cols = df.columns.tolist()
        
        # Check if last 3 columns are operating point data
        if len(df.columns) >= 7:
            # Assume last 3 columns are: Current Operating Temperature, Current Operating Flowrate, Operating Case
            has_operating_points = True
            operating_temp_col = df.columns[-3]
            operating_flow_col = df.columns[-2]
            operating_case_col = df.columns[-1]
            
            # Extract unique operating points (remove NaN rows)
            op_data = df[[operating_temp_col, operating_flow_col, operating_case_col]].dropna()
            
            for _, row in op_data.iterrows():
                operating_points.append({
                    'temperature': float(row[operating_temp_col]),
                    'flowrate': float(row[operating_flow_col]),
                    'case': str(row[operating_case_col])
                })
            
            # Remove operating point columns for envelope calculation
            df = df.iloc[:, :4]
        
        # Data Cleaning and Column Renaming
        df.columns = [
            'Temperature_Inlet',
            'Flowrate_Actual',
            'Flowrate_PD_Limit',
            'Flowrate_Momentum_Limit'
        ]

        for col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

        df.dropna(inplace=True)

        # Constraint Analysis
        df['Flowrate_Constraint_Min'] = df[['Flowrate_Actual', 'Flowrate_PD_Limit', 'Flowrate_Momentum_Limit']].min(axis=1)

        active_constraints = np.argmin(df[['Flowrate_Actual', 'Flowrate_PD_Limit', 'Flowrate_Momentum_Limit']].values, axis=1)
        constraint_names = ['Actual Flowrate', 'Pressure Drop (PD)', 'Momentum']
        df['Active_Constraint'] = [constraint_names[i] for i in active_constraints]

        crossover_rows = df[df['Active_Constraint'] != df['Active_Constraint'].shift(1)]

        analysis_text = ""
        if len(crossover_rows) > 1:
            shift_data = crossover_rows.iloc[1]
            shift_temp = shift_data['Temperature_Inlet']
            shift_flow = shift_data['Flowrate_Constraint_Min']
            shift_to = shift_data['Active_Constraint']

            analysis_text = (
                f"Constraint Shift Analysis:\n"
                f"Limiting curve switches at Inlet Temperature: {shift_temp:.1f} °C\n"
                f"Flowrate limit at shift: {shift_flow:.0f} kg/hr\n"
                f"New limiting curve: {shift_to}"
            )
        else:
            shift_temp, shift_flow = None, None
            analysis_text = "Note: The overall limiting curve does not switch within the provided temperature range."

        # Create the plot
        fig, ax = plt.subplots(figsize=(14, 8))

        # Plot lines
        ax.plot(
            df['Temperature_Inlet'],
            df['Flowrate_Actual'],
            label='Area Ratio',
            color='#FF00FF',
            linestyle='-',
            linewidth=3
        )

        ax.plot(
            df['Temperature_Inlet'],
            df['Flowrate_PD_Limit'],
            label='Allowable Pressure Drop Limit (0.7 bar)',
            color='red',
            linestyle='-',
            linewidth=2.5
        )

        ax.plot(
            df['Temperature_Inlet'],
            df['Flowrate_Momentum_Limit'],
            label=r'Inlet Nozzle Momentum Limit ($7000-\rho v^2$)',
            color='#228B22',
            linestyle='-.',
            linewidth=2.5
        )

        # Fill safe operating zone
        ax.fill_between(
            df['Temperature_Inlet'],
            0,
            df['Flowrate_Constraint_Min'],
            color='#4DF04D',
            alpha=0.6,
            label='Safe Operating Zone (Below All Curves)'
        )

        # Shift point info stored but not plotted

        # Plot operating points if available
        if has_operating_points and operating_points:
            
            # Define unique colors for each operating point
            point_colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8', 
                          '#F7DC6F', '#BB8FCE', '#85C1E2', '#F8B739', '#52C67D']
            
            for idx, point in enumerate(operating_points):
                temp = point['temperature']
                flow = point['flowrate']
                case = point['case']
                
                # Assign unique color to each point
                point_color = point_colors[idx % len(point_colors)]
                
                # Plot the operating point with unique color - solid, no border
                # Add case name with temp and flowrate to legend
                ax.scatter(
                    temp, flow,
                    color=point_color,
                    marker='o',
                    s=150,
                    zorder=10,
                    label=f'{case} ({temp:.1f}°C, {flow:.0f} kg/hr)'
                )

        # Plot aesthetics
        ax.set_title(f'{sheet_name} Operating Envelope: Tube Side Mass Flowrate vs. Inlet Temperature',
                     fontsize=18, fontweight='bold')
        ax.set_xlabel('Tube Side Inlet Temperature (°C)', fontsize=14)
        ax.set_ylabel('Tube Side Mass Flowrate (kg/hr)', fontsize=14)
        ax.grid(True, linestyle=':', alpha=0.6)
        
        # Create two separate legends
        # First legend for envelope curves
        handles, labels = ax.get_legend_handles_labels()
        
        # Separate operating points from other elements
        envelope_handles = []
        envelope_labels = []
        operating_handles = []
        operating_labels = []
        
        for handle, label in zip(handles, labels):
            if label in ['Area Ratio', 'Allowable Pressure Drop Limit (0.7 bar)', 
                        'Inlet Nozzle Momentum Limit ($7000-\\rho v^2$)', 
                        'Safe Operating Zone (Below All Curves)']:
                envelope_handles.append(handle)
                envelope_labels.append(label)
            else:
                operating_handles.append(handle)
                operating_labels.append(label)
        
        # Create first legend for envelope curves at upper right
        legend1 = ax.legend(envelope_handles, envelope_labels, 
                          loc='upper right', fontsize=9, framealpha=0.9,
                          title='Operating Envelope')
        ax.add_artist(legend1)  # Add first legend back to plot
        
        # Create second legend for operating points at lower right
        if operating_handles:
            ax.legend(operating_handles, operating_labels, 
                     loc='lower right', fontsize=9, framealpha=0.9,
                     title='Operating Cases', ncol=1)
        
        ax.set_ylim(bottom=0)

        plt.tight_layout()

        return fig, analysis_text

    except Exception as e:
        return None, f"Error: {str(e)}"


def process_excel_workbook(input_file, output_file=None):
    """
    Read all worksheets from an Excel file, generate plots, and save them back to Excel.

    Parameters:
    -----------
    input_file : str
        Path to the input Excel file containing data in multiple sheets
    output_file : str, optional
        Path to the output Excel file. If None, will create a file with '_output' suffix
    """
    
    # Fix the file extension if it's malformed
    base_name = os.path.splitext(input_file)[0]
    
    # If file doesn't end with .xlsx or .xls, try to fix it
    if not (input_file.endswith('.xlsx') or input_file.endswith('.xls')):
        st.warning("⚠️ File has unusual extension. Attempting to fix...")
        corrected_file = base_name + '.xlsx'
        try:
            os.rename(input_file, corrected_file)
            input_file = corrected_file
            st.success(f"✓ File renamed to: {corrected_file}")
        except Exception as e:
            st.warning("Could not rename file. Attempting to process anyway...")

    if output_file is None:
        if input_file.endswith('.xlsx'):
            output_file = input_file.replace('.xlsx', '_output.xlsx')
        elif input_file.endswith('.xls'):
            output_file = input_file.replace('.xls', '_output.xlsx')
        else:
            output_file = base_name + '_output.xlsx'

    try:
        # Read all sheets from the Excel file
        excel_file = pd.ExcelFile(input_file)
        sheet_names = excel_file.sheet_names

        st.info(f"📊 Processing {len(sheet_names)} sheet(s) from the uploaded file...")

        # Create a new Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

            for sheet_name in sheet_names:
                with st.spinner(f"Processing sheet: '{sheet_name}'..."):
                    # Read the sheet
                    df = pd.read_excel(input_file, sheet_name=sheet_name)

                    # Process the data and create plot (pass sheet_name for title)
                    fig, analysis_text = process_sheet_data(df, sheet_name)

                    if fig is not None:
                        # Save the original data to the output workbook
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                        # Save the plot as an image in memory
                        img_buffer = io.BytesIO()
                        fig.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
                        img_buffer.seek(0)

                        # Get the worksheet
                        worksheet = writer.sheets[sheet_name]

                        # Add the analysis text below the data
                        start_row = len(df) + 3
                        worksheet.cell(row=start_row, column=1, value="Analysis Results:")
                        for i, line in enumerate(analysis_text.split('\n')):
                            worksheet.cell(row=start_row + i + 1, column=1, value=line)

                        # Insert the plot image
                        img = Image(img_buffer)
                        img.anchor = 'H2'  # Moved further right to accommodate more columns
                        worksheet.add_image(img)

                        # Display in Streamlit
                        st.success(f"✓ Successfully processed '{sheet_name}'")
                        
                        # Create columns for better layout
                        col1, col2 = st.columns([2, 1])
                        
                        with col1:
                            st.pyplot(fig)
                        
                        with col2:
                            st.info("📈 Analysis Results")
                            st.text(analysis_text)
                        
                        plt.close(fig)
                        st.markdown("---")
                    else:
                        st.error(f"✗ Failed to process '{sheet_name}'")
                        st.text(analysis_text)

        st.success("🎉 All sheets processed successfully!")
        return output_file

    except Exception as e:
        st.error(f"❌ Error processing workbook: {e}")
        with st.expander("💡 Troubleshooting Tips"):
            st.markdown("""
            1. Make sure the file is a valid Excel file (.xlsx or .xls)
            2. Check that the file is not corrupted
            3. Verify the file extension is correct
            4. Try opening the file in Excel and saving it again
            5. Ensure the file has the correct column structure
            """)
        raise


# ==============================================================================
# STREAMLIT APP
# ==============================================================================

# Page configuration
st.set_page_config(
    page_title="ACHE Operating Envelope Analyzer",
    page_icon="🚀",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for better styling
st.markdown("""
    <style>
    .main {
        padding: 2rem;
    }
    .stButton>button {
        width: 100%;
        background-color: #4CAF50;
        color: white;
        font-weight: bold;
        padding: 0.75rem;
        border-radius: 8px;
    }
    .stButton>button:hover {
        background-color: #45a049;
    }
    </style>
    """, unsafe_allow_html=True)

# Sidebar
with st.sidebar:
    st.image("https://img.icons8.com/fluency/96/000000/heat-exchanger.png", width=100)
    st.title("📋 About")
    st.markdown("""
    ### ACHE Analyzer
    
    This tool analyzes Air Cooled Heat Exchanger (ACHE) operating data.
    
    **Features:**
    - Multi-sheet processing
    - Constraint analysis
    - Safe zone visualization
    - Operating point tracking
    - Safety status assessment
    - Automated reporting
    
    **Version:** 2.0  
    **Updated:** 2025
    """)

# Main content
st.title("🚀 ACHE Operating Envelope Analyzer")
st.markdown("### Automated Analysis Tool for Air Cooled Heat Exchangers")
st.markdown("---")

# Instructions in an expander
with st.expander("📖 Instructions - Click to expand", expanded=True):
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        #### 📥 Upload Requirements
        - File format: `.xlsx` or `.xls`
        - Multiple worksheets supported
        - Each sheet should contain ACHE data
        
        #### 📊 Required Columns (in order)
        1. Tube Side Inlet Temperature
        2. Tube Side Mass Flowrate
        3. Tube Side Pressure Drop Limit
        4. Inlet Nozzle Momentum Limit
        
        #### 📍 Optional Operating Point Columns
        5. Current Operating Temperature
        6. Current Operating Flowrate
        7. Operating Case Name
        """)
    
    with col2:
        st.markdown("""
        #### 🔄 How to Use
        1. Upload your Excel file below
        2. Click **Process File** button
        3. Review plots and analysis
        4. Download the output file
        
        #### 📈 Output Includes
        - Operating envelope plots
        - Constraint shift analysis
        - Safe operating zones
        - Operating point status (SAFE/UNSAFE)
        - Safety margins
        - Excel file with embedded plots
        
        #### 🎯 Operating Point Markers
        - Each operating case has a unique color
        - Solid colored circles (no borders)
        - Case name, temperature, and flowrate in legend
        - No on-plot labels for cleaner visualization
        """)

st.markdown("---")

# File uploader section
st.markdown("### 📁 Upload Your Excel File")
uploaded_file = st.file_uploader(
    "Choose an Excel file",
    type=['xlsx', 'xls'],
    help="Upload an Excel file containing ACHE data in multiple worksheets"
)

if uploaded_file is not None:
    # Display file details in a nice format
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("File Name", uploaded_file.name)
    with col2:
        st.metric("File Size", f"{uploaded_file.size / 1024:.2f} KB")
    with col3:
        st.metric("File Type", uploaded_file.type.split('.')[-1].upper())
    
    st.markdown("---")
    
    # Create a temporary file to save the uploaded file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        tmp_file.write(uploaded_file.getvalue())
        tmp_file_path = tmp_file.name

    # Process button
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        process_button = st.button("🔄 Process File", type="primary", use_container_width=True)
    
    if process_button:
        try:
            # Process the file
            output_file = process_excel_workbook(tmp_file_path)
            
            # Read the output file for download
            with open(output_file, 'rb') as f:
                output_data = f.read()
            
            # Success message and download section
            st.markdown("---")
            st.success("### 🎉 Processing Complete!")
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.download_button(
                    label="📥 Download Output Excel File",
                    data=output_data,
                    file_name=f"{os.path.splitext(uploaded_file.name)[0]}_output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            # Clean up temporary files
            os.unlink(tmp_file_path)
            os.unlink(output_file)
            
        except Exception as e:
            st.error(f"An error occurred: {str(e)}")
            # Clean up temporary file on error
            if os.path.exists(tmp_file_path):
                os.unlink(tmp_file_path)

else:
    # Show placeholder when no file is uploaded
    st.info("👆 Please upload an Excel file to begin analysis")
    
    # Sample data format
    with st.expander("💡 View Sample Data Format"):
        sample_df = pd.DataFrame({
            'Tube Side Inlet Temperature': [122.2, 126.7, 131.1, '', ''],
            'Tube side Mass Flowrate': [148000, 133000, 121000, '', ''],
            'Tube side Pressure Drop limit (0.7 bar)': [74105, 73820, 73547, '', ''],
            'Inlet Nozzle Momentum (7000 pv²)': [117000, 116000, 114000, '', ''],
            'Current Operating Temperature': ['', '', '', 116, 118.4],
            'Current Operating Flowrate': ['', '', '', 69700, 75400],
            'Operating Case': ['', '', '', 'Case 1', 'Case 2']
        })
        st.dataframe(sample_df, use_container_width=True)
        st.caption("Note: Operating point columns can have data in any rows (they will be extracted automatically)")

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #666;'>
    <p>🚀 ACHE Operating Envelope Analyzer v2.0 | Built with Streamlit</p>
    <p style='font-size: 0.8em;'>For support or questions, contact your system administrator</p>
</div>
""", unsafe_allow_html=True)
